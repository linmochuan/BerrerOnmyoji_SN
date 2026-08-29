"""lin_RPA 启动器。"""
from __future__ import annotations

from dataclasses import replace
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from rpa_executor import ExcelExecutor, start_async
from rpa_operations import DesktopOperations, StopController
from rpa_readers import AppConfig, read_app_config, read_model_metadata, save_settings, validate_startup_inputs


class LinRPAApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("lin RPA")
        self.root.geometry("760x540")
        self.root.minsize(680, 480)
        self.stop_controller = StopController()
        self.listener = None
        self.worker = None
        self._last_log_message = None
        self._last_log_line = None
        self.config = read_app_config()
        self._build_ui()
        self._load_values()
        self._start_global_hotkey()
        self.root.protocol("WM_DELETE_WINDOW", self.close)

    def _build_ui(self) -> None:
        frame = ttk.Frame(self.root, padding=16)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="lin RPA", font=("Microsoft YaHei", 18, "bold")).pack(anchor=tk.W)
       

        self.excel_var = tk.StringVar()
        self.sheet_var = tk.StringVar()
        self.model_var = tk.StringVar()
        self.screenshot_var = tk.StringVar()
        self.confidence_var = tk.StringVar(value=str(self.config.confidence))
        self.duration_var = tk.StringVar(value="0")

        self._path_row(frame, "Excel 文件", self.excel_var, self.choose_excel)
        self._path_row(frame, "模型文件", self.model_var, self.choose_model)
        sheet_row = ttk.Frame(frame)
        sheet_row.pack(fill=tk.X, pady=5)
        ttk.Label(sheet_row, text="工作表", width=12).pack(side=tk.LEFT)
        self.sheet_box = ttk.Combobox(sheet_row, textvariable=self.sheet_var, state="readonly")
        self.sheet_box.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(sheet_row, text="刷新", command=self.refresh_sheets).pack(side=tk.LEFT, padx=(6, 0))
        self._path_row(frame, "截图保存目录", self.screenshot_var, self.choose_screenshot)

        settings = ttk.Frame(frame)
        settings.pack(fill=tk.X, pady=5)
        ttk.Label(settings, text="置信度", width=12).pack(side=tk.LEFT)
        ttk.Entry(settings, textvariable=self.confidence_var, width=10).pack(side=tk.LEFT)
        ttk.Label(settings, text="运行分钟（0=持续）", width=18).pack(side=tk.LEFT, padx=(20, 0))
        ttk.Entry(settings, textvariable=self.duration_var, width=10).pack(side=tk.LEFT)

        buttons = ttk.Frame(frame)
        buttons.pack(fill=tk.X, pady=(14, 10))
        ttk.Button(buttons, text="开始", command=self.start_excel).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(buttons, text="停止", command=self.stop).pack(side=tk.LEFT)

        ttk.Label(frame, text="运行日志").pack(anchor=tk.W)
        log_frame = ttk.Frame(frame)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(4, 0))
        self.log = tk.Text(log_frame, height=12, state=tk.DISABLED, wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log.yview)
        self.log.configure(yscrollcommand=scrollbar.set)
        self.log.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def _path_row(self, parent, label: str, variable: tk.StringVar, command) -> None:
        row = ttk.Frame(parent)
        row.pack(fill=tk.X, pady=5)
        ttk.Label(row, text=label, width=12).pack(side=tk.LEFT)
        ttk.Entry(row, textvariable=variable).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(row, text="选择", command=command).pack(side=tk.LEFT, padx=(6, 0))

    def _load_values(self) -> None:
        self.excel_var.set(self.config.excel_path)
        self.model_var.set(self.config.model_path)
        self.screenshot_var.set(self.config.screenshot_path)
        self.refresh_sheets()

    def choose_excel(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx;*.xls")])
        if path:
            self.excel_var.set(path)
            self.refresh_sheets()

    def choose_model(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("PyTorch 模型", "*.pt;*.pth"), ("所有文件", "*.*")])
        if path:
            self.model_var.set(path)
            self._sync_model_config()

    def _sync_model_config(self) -> None:
        model_path = self.model_var.get().strip()
        if not model_path:
            return
        class_names, translations = read_model_metadata(model_path, self.config.root)
        self.config = replace(
            self.config,
            model_path=model_path,
            class_names=class_names,
            class_name_to_chinese=translations,
        )

    def choose_screenshot(self) -> None:
        path = filedialog.askdirectory()
        if path:
            self.screenshot_var.set(path)

    def refresh_sheets(self) -> None:
        path = self.excel_var.get().strip()
        if not path or not Path(path).exists():
            return
        import openpyxl
        workbook = openpyxl.load_workbook(path, read_only=True)
        try:
            self.sheet_box["values"] = workbook.sheetnames
            if workbook.sheetnames and not self.sheet_var.get():
                self.sheet_var.set(workbook.sheetnames[0])
        finally:
            workbook.close()

    def _make_config(self) -> AppConfig:
        confidence = min(1.0, max(0.1, float(self.confidence_var.get())))
        model_path = self.model_var.get().strip()
        class_names, translations = read_model_metadata(model_path, self.config.root) if model_path else (self.config.class_names, self.config.class_name_to_chinese)
        config = AppConfig(
            root=self.config.root,
            confidence=confidence,
            excel_path=self.excel_var.get().strip(),
            sheet_name=self.sheet_var.get().strip(),
            screenshot_path=self.screenshot_var.get().strip(),
            hotkey=self.config.hotkey,
            model_path=model_path,
            class_names=class_names,
            class_name_to_chinese=translations,
            actions=self.config.actions,
        )
        self.config = config
        save_settings(config)
        return config

    def _log(self, message: str) -> None:
        self.root.after(0, self._append_log, message)

    def _append_log(self, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M")
        if message == self._last_log_message and self._last_log_line is not None:
            self.log.configure(state=tk.NORMAL)
            self.log.delete(f"{self._last_log_line}.0", f"{self._last_log_line}.end")
            self.log.insert(f"{self._last_log_line}.0", f"{self._last_log_line_text}【{timestamp}】")
            self.log.configure(state=tk.DISABLED)
            self.log.see(tk.END)
            return

        self._last_log_message = message
        self._last_log_line = int(self.log.index("end-1c").split(".")[0])
        self._last_log_line_text = f"{timestamp}：{message}"
        self.log.configure(state=tk.NORMAL)
        self.log.insert(tk.END, f"{self._last_log_line_text}\n")
        self.log.see(tk.END)
        self.log.configure(state=tk.DISABLED)

    def start_excel(self) -> None:
        try:
            if self.worker is not None and self.worker.is_alive():
                messagebox.showwarning("已运行", "当前任务已经在执行中，请先停止后再启动。")
                return
            config = self._make_config()
            validate_startup_inputs(config)
            duration = int(self.duration_var.get() or 0) * 60
            self.stop_controller.reset()
            operations = DesktopOperations(self.stop_controller, self._log)
            executor = ExcelExecutor(config, operations, self._log)
            self.worker = start_async(lambda: executor.run(config.excel_path, config.sheet_name, duration))
            self._log("已启动")
        except (OSError, ValueError, TypeError) as error:
            messagebox.showerror("启动失败", str(error))

    def _start_global_hotkey(self) -> None:
        if self.config.hotkey:
            operations = DesktopOperations(self.stop_controller, self._log)
            self.listener = operations.start_hotkey_listener(set(self.config.hotkey), self.stop)

    def stop(self) -> None:
        self.stop_controller.stop()
        self._log("已发送停止信号")

    def close(self) -> None:
        self.stop()
        if self.listener:
            self.listener.stop()
        self.root.destroy()


def main() -> None:
    root = tk.Tk()
    LinRPAApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
