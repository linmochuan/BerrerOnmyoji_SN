import openpyxl
import pyautogui
import time
import datetime
import os
import random
import threading
import numpy as np
import tkinter as tk
import webbrowser
import json
from tkinter import filedialog
from tkinter import ttk
import pandas as pd
import tkinter.messagebox as messagebox
from pynput import keyboard

def click (Shiji_x, Shiji_y):
    pyautogui.moveTo(Shiji_x, Shiji_y)
    pyautogui.mouseDown()
    time.sleep(random.uniform(0.06464266777038574, 0.1))# 点击时长
    pyautogui.mouseUp()

def generate_random_position(center, min_val, max_val, size=1):
    """生成指定范围内的随机位置"""
    std_dev = (max_val - min_val) / 6.0
    values = np.random.normal(loc=center, scale=std_dev, size=size)
    return np.clip(values, min_val, max_val)

class AutoScriptApp:
    def __init__(self, root):
        """初始化应用程序"""
        try:
            # 设置错误日志路径
            self.log_folder = "错误喵"
            if not os.path.exists(self.log_folder):
                os.makedirs(self.log_folder)
            
            # 记录启动时间
            self.start_time = datetime.datetime.now()
            self.log_error(f"程序启动时间: {self.start_time}")
            
            if not self.check_program_validity():
                self.log_error("程序验证失败")
                root.destroy()
                return
            
            self.root = root
            self.root.title("龙钰的护肝小助手")
            
            # 设置窗口图标
            try:
                self.root.iconbitmap('pro.ico')
            except Exception as e:
                self.log_error(f"加载图标失败: {str(e)}")
            
            # 配置文件设置
            self.CONFIG_FILE = "settings.json"
            self.log_error(f"配置文件路径: {os.path.abspath(self.CONFIG_FILE)}")
            
            # 初始化变量
            self.init_variables()
            
            # 加载配置
            self.load_settings()
            
            # 初始化UI
            self.init_ui()
            
            # 加载快捷键
            self.load_hotkey()
            
            # 启动快捷键监听
            self.start_hotkey_listener()
            
        except Exception as e:
            self.log_error(f"初始化过程出错: {str(e)}\n{self.get_traceback()}")
            raise
    
    def init_variables(self):
        """初始化所有必要的变量"""
        # 设置默认值
        self.DEFAULT_CONFIDENCE = 0.8
        self.DEFAULT_RUN_TIME = 60
        self.DEFAULT_CUSTOM_TIME = None
        
        # 创建日志文件夹
        self.log_folder = "错误喵"
        if not os.path.exists(self.log_folder):
            os.makedirs(self.log_folder)
        
        self.UPDATE_INTERVAL = 1000
        self.is_running = False
        self.stop_signal = False
        
        # 添加Excel文件和工作表相关变量
        self.excel_file = None
        self.selected_sheet = None
        self.excel_path = ""
        self.screenshot_path = ""
        
        # 初始化其他必要的变量
        self.confidence_entry = None
        self.hotkey = set()
        self.current_keys = set()
        self.listener = None
        self.is_setting_hotkey = False
        self.temp_hotkey = set()
    
    def init_ui(self):
        """将UI初始化代码单独提取为方法"""
        try:
            # 主内容框架
            main_frame = tk.Frame(self.root)
            main_frame.pack(pady=(50, 5))
            
            # 将日志区域的创建移到最前面
            # 创建日志区域
            self.log = tk.Text(main_frame, height=15, width=60)
            self.log.pack(pady=5)
            
            # Excel文件选择框架
            excel_frame = tk.Frame(main_frame)
            excel_frame.pack(pady=5)
            
            self.excel_label = tk.Label(excel_frame, text="Excel文件路径：")
            self.excel_label.pack(side=tk.LEFT)
            self.excel_entry = tk.Entry(excel_frame, width=30)
            self.excel_entry.pack(side=tk.LEFT, padx=5)
            tk.Button(excel_frame, text="选择文件", 
                     command=self.select_excel_file).pack(side=tk.LEFT)

            # 工作表选择框架
            sheet_frame = tk.Frame(main_frame)
            sheet_frame.pack(pady=5)
            
            self.sheet_label = tk.Label(sheet_frame, text="工作表名称：")
            self.sheet_label.pack(side=tk.LEFT)
            self.sheet_combobox = ttk.Combobox(sheet_frame, width=27, state='readonly')
            self.sheet_combobox.pack(side=tk.LEFT, padx=5)
            tk.Button(sheet_frame, text="刷新", 
                     command=self.refresh_sheets).pack(side=tk.LEFT)

            # 截图保存位置选择框架
            screenshot_frame = tk.Frame(main_frame)
            screenshot_frame.pack(pady=5)
            
            self.screenshot_label = tk.Label(screenshot_frame, text="截图保存位置：")
            self.screenshot_label.pack(side=tk.LEFT)
            self.screenshot_entry = tk.Entry(screenshot_frame, width=30)
            self.screenshot_entry.pack(side=tk.LEFT, padx=5)
            tk.Button(screenshot_frame, text="选择文件夹", 
                     command=self.choose_screenshot_path).pack(side=tk.LEFT)

            # 运行时间框架
            time_frame = tk.Frame(main_frame)
            time_frame.pack(pady=5)
            
            # 自定义时间输入
            custom_frame = tk.Frame(time_frame)
            custom_frame.pack(pady=5)
            self.custom_time_label = tk.Label(custom_frame, text="运行时间（分钟）：")
            self.custom_time_label.pack(side=tk.LEFT)
            self.custom_time_entry = tk.Entry(custom_frame, width=10)
            self.custom_time_entry.pack(side=tk.LEFT, padx=5)
            
            # 在创建完custom_time_entry后立即设置值
            if hasattr(self, 'DEFAULT_RUN_TIME'):
                self.custom_time_entry.insert(0, str(self.DEFAULT_RUN_TIME))
            
            # 图像匹配精度输入框
            confidence_frame = tk.Frame(main_frame)
            confidence_frame.pack(pady=5)
            self.confidence_label = tk.Label(confidence_frame, text="图像匹配精度（0.1-1.0）：")
            self.confidence_label.pack(side=tk.LEFT)
            self.confidence_entry = tk.Entry(confidence_frame, width=10)
            self.confidence_entry.insert(0, str(self.DEFAULT_CONFIDENCE))
            self.confidence_entry.pack(side=tk.LEFT, padx=5)

            # 控制按钮
            btn_frame = tk.Frame(main_frame)
            btn_frame.pack(pady=5)
            
            tk.Button(btn_frame, text="开始", command=self.start).pack(side=tk.LEFT, padx=5)
            tk.Button(btn_frame, text="停止", command=self.stop).pack(side=tk.LEFT, padx=5)

            # 不要点我按钮
            self.open_web_button = tk.Button(self.root, text="不要点我", 
                                            command=self.open_webpage)
            self.open_web_button.place(x=10, y=10)

            # 加载上次设置
            if hasattr(self, 'excel_path') and self.excel_path:
                self.excel_entry.insert(0, self.excel_path)
                self.refresh_sheets()
                
            if hasattr(self, 'screenshot_path') and self.screenshot_path:
                self.screenshot_entry.insert(0, self.screenshot_path)
                
            if hasattr(self, 'DEFAULT_CUSTOM_TIME') and self.DEFAULT_CUSTOM_TIME:
                self.custom_time_entry.insert(0, str(self.DEFAULT_CUSTOM_TIME))

            self.update_clock()
            self.root.protocol("WM_DELETE_WINDOW", self.on_close)
            
            # 添加快捷键停止相关控件
            self.hotkey_frame = ttk.Frame(self.root)
            self.hotkey_frame.pack(pady=5)
            
            self.hotkey_button = ttk.Button(self.hotkey_frame, text="开始设置快捷键停止", 
                                           command=self.start_hotkey_setting)
            self.hotkey_button.pack(side=tk.LEFT, padx=5)
            
            self.hotkey_display = ttk.Entry(self.hotkey_frame, width=20, state='readonly')
            self.hotkey_display.pack(side=tk.LEFT, padx=5)
            
            # 更新快捷键显示
            if hasattr(self, 'hotkey'):
                self.update_hotkey_display()
            
        except Exception as e:
            self.log_error(f"初始化UI时出错: {str(e)}\n{self.get_traceback()}")
    
    def create_excel_selection(self):
        """创建Excel文件选择相关控件"""
        excel_frame = tk.Frame(self.root)
        excel_frame.pack(pady=5)
        
        # Excel文件选择
        self.excel_label = tk.Label(excel_frame, text="Excel文件路径：")
        self.excel_label.pack(side=tk.LEFT)
        self.excel_entry = tk.Entry(excel_frame, width=30)
        self.excel_entry.pack(side=tk.LEFT, padx=5)
        tk.Button(excel_frame, text="选择文件", 
                 command=self.select_excel_file).pack(side=tk.LEFT)
        
        # 工作表选择下拉框
        self.sheet_var = tk.StringVar()
        self.sheet_combobox = ttk.Combobox(excel_frame, 
                                          textvariable=self.sheet_var,
                                          state='readonly')
        self.sheet_combobox.pack()

        # 添加截图保存位置选择
        screenshot_frame = tk.Frame(self.root)
        screenshot_frame.pack(pady=5)
        
        self.screenshot_label = tk.Label(screenshot_frame, text="截图保存位置：")
        self.screenshot_label.pack(side=tk.LEFT)
        self.screenshot_entry = tk.Entry(screenshot_frame, width=30)
        self.screenshot_entry.pack(side=tk.LEFT, padx=5)
        tk.Button(screenshot_frame, text="选择文件夹", 
                 command=self.choose_screenshot_path).pack(side=tk.LEFT)
    
    def select_excel_file(self):
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx;*.xls")]
        )
        
        if file_path:
            self.excel_path = file_path
            self.excel_entry.delete(0, tk.END)
            self.excel_entry.insert(0, file_path)
            self.refresh_sheets()
            return file_path
        return None

    def refresh_sheets(self):
        """刷新工作表列表"""
        if self.excel_path and os.path.exists(self.excel_path):
            try:
                wb = openpyxl.load_workbook(self.excel_path)
                self.sheet_combobox['values'] = wb.sheetnames
                if not self.sheet_combobox.get() and wb.sheetnames:
                    self.sheet_combobox.set(wb.sheetnames[0])
                wb.close()
            except Exception as e:
                self.log_error(f"读取工作表失败喵~ {str(e)}")

    def choose_screenshot_path(self):
        """选择截图保存位置"""
        folder_path = filedialog.askdirectory(
            title="选择截图保存位置"
        )
        if folder_path:
            self.screenshot_path = folder_path
            self.screenshot_entry.delete(0, tk.END)
            self.screenshot_entry.insert(0, folder_path)

    def update_listbox(self):
        # 清空现有列表
        self.listbox.delete(0, tk.END)
        
        # 如果excel_data存在，显示数据
        global excel_data
        if excel_data is not None:
            # 根据您的Excel结构调整相应的列名
            for index, row in excel_data.iterrows():
                # 这里假设Excel中有'名称'列，请根据实际情况修改
                self.listbox.insert(tk.END, str(row))  
    
    def create_time_selection(self):
        """创建时间选择相关控件"""
        time_frame = tk.Frame(self.root)
        time_frame.pack(pady=5)
        
        tk.Label(time_frame, text="选择运行时间:").pack()
        
        self.run_time_var = tk.IntVar(value=0 if self.DEFAULT_CUSTOM_TIME else self.DEFAULT_RUN_TIME)
        times = [(10, "10分钟"), (30, "30分钟"), (60, "1小时"), (0, "自定义(分钟)")]
        
        for value, text in times:
            tk.Radiobutton(time_frame, text=text, 
                          variable=self.run_time_var, 
                          value=value).pack()
        
        self.custom_time_entry = tk.Entry(time_frame)
        self.custom_time_entry.pack()
        
        # 如果有保存的自定义时间，则填入
        if self.DEFAULT_CUSTOM_TIME:
            self.custom_time_entry.insert(0, str(self.DEFAULT_CUSTOM_TIME))
    
    def create_confidence_setting(self):
        """创建精度设置相关控件"""
        conf_frame = tk.Frame(self.root)
        conf_frame.pack(pady=5)
        
        tk.Label(conf_frame, text="设置图像匹配精度（0.1-1.0）:").pack()
        self.confidence_entry = tk.Entry(conf_frame)
        # 使用加载的配置值
        self.confidence_entry.insert(0, str(self.DEFAULT_CONFIDENCE))
        self.confidence_entry.pack()
        
        # 添加配置值变化监听
        self.confidence_entry.bind('<FocusOut>', self.on_confidence_change)
    
    def create_control_buttons(self):
        """创建控制按钮"""
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=5)
        
        tk.Button(btn_frame, text="开始", command=self.start).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="停止", command=self.stop).pack(side=tk.LEFT, padx=5)
        
        # 将网页按钮移到更合适的位置
        self.open_web_button = tk.Button(self.root, text="不要点我", 
                                        command=self.open_webpage)
        self.open_web_button.place(x=10, y=10)
    
    def create_log_area(self):
        """创建日志区域"""
        self.log = tk.Text(self.root, height=15, width=60)
        self.log.pack(pady=5)
    
    def open_webpage(self):
        webbrowser.open("https://m.tb.cn/h.TfSvhID?tk=FXag3CXbMRA")
    
    def update_clock(self):
        """更新时钟显示"""
        now = time.strftime("%Y-%m-%d %H:%M:%S")
        self.root.title(f"龙钰的护肝小助手 - 当前时间: {now}")
        self.root.after(1000, self.update_clock)
    
    def log_message(self, message):
        """添加安全检查"""
        if hasattr(self, 'log') and self.log is not None:
            self.log.insert(tk.END, f"{message}\n")
            self.log.see(tk.END)
        else:
            print(f"日志组件未初始化: {message}")  # 降级到控制台输出
    
    def start(self):
        """开始执行程序"""
        try:
            # 检查Excel文件是否已选择并且存在
            excel_path = self.excel_entry.get().strip()
            if not excel_path or not os.path.exists(excel_path):
                self.log_message("请先选择Excel文件喵~")
                return
            
            # 检查是否选择了工作表
            sheet_name = self.sheet_combobox.get()
            if not sheet_name:
                self.log_message("请选择工作表喵~")
                return

            try:
                # 尝试打开Excel文件验证其可用性
                wb = openpyxl.load_workbook(excel_path)
                if sheet_name not in wb.sheetnames:
                    self.log_message("选择的工作表不存在喵~")
                    wb.close()
                    return
                wb.close()
            except Exception as e:
                self.log_message(f"打开Excel文件失败喵~ {str(e)}")
                return

            self.log_message(f"{datetime.datetime.now()}程序启动...")
            self.stop_signal = False
            
            # 获取运行时间
            custom_time = self.custom_time_entry.get().strip()
            if not custom_time:
                self.log_message("请输入运行时间（分钟）")
                return
            
            try:
                run_time = int(custom_time) * 60  # 转换为秒
            except ValueError:
                self.log_message("请输入有效的运行时间（分钟）")
                return
            
            # 获取图像匹配精度
            try:
                self.confidence = float(self.confidence_entry.get())
                if not (0.1 <= self.confidence <= 1.0):
                    raise ValueError
                # 保存当前配置
                self.save_settings()
            except ValueError:
                self.log_message(f"无效的精度值，使用默认值 {self.DEFAULT_CONFIDENCE}")
                self.confidence = self.DEFAULT_CONFIDENCE
            
            # 更新Excel文件路径和工作表名称
            self.excel_path = excel_path
            self.sheet_name = sheet_name
            
            self.start_time = time.time()
            self.end_time = self.start_time + run_time
            
            # 显示预计结束时间
            end_datetime = datetime.datetime.now() + datetime.timedelta(seconds=run_time)
            self.log_message(f"预计在 {end_datetime.strftime('%H:%M:%S')} 完成任务喵~")
            
            self.is_running = True
            self.thread = threading.Thread(target=self.run_script)
            self.thread.start()
            
        except Exception as e:
            self.log_error(f"启动程序时出错: {str(e)}\n{self.get_traceback()}")
    
    def stop(self):
        self.log_message(f"{datetime.datetime.now().strftime('%H:%M:%S')}程序停止...")
        self.is_running = False
        self.stop_signal = True
    
    def on_close(self):
        if self.listener:
            self.listener.stop()
        self.stop()
        self.root.destroy()
    
    def run_script(self):
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            sheet = wb[self.sheet_name]
            
            max_row = sheet.max_row
            
            if max_row < 2:
                self.log_message("主人，Excel文件是空的喵~")
                return
                
            current_row = 2
            while self.is_running and time.time() < self.end_time:
                try:
                    # 检查行号是否有效
                    if current_row > max_row:
                        self.log_message("主人，已经读完表格了喵，重新从第2行开始咯~")
                        current_row = 2
                        continue
                    
                    row_data = list(sheet.iter_rows(min_row=current_row, max_row=current_row, 
                                                  max_col=4, values_only=True))[0]
                    image_name, operations, search_timeout, timeout_action = row_data

                    if not image_name or not operations:
                        current_row += 1
                        continue

                    # 读取图片路径
                    image_path = os.path.join(image_name)

                    # 设置查找超时时间
                    search_start_time = time.time()
                    image_found = False
                    
                    # 在超时时间内尝试查找图片
                    while time.time() - search_start_time < float(search_timeout or 10):  # 默认10秒
                        try:
                            location = pyautogui.locateCenterOnScreen(image_path, confidence=self.confidence)
                            image_location = pyautogui.locateOnScreen(image_path, confidence=self.confidence)
                            if location and image_location:
                                image_found = True
                                self.log_message(
                                    f"{datetime.datetime.now().strftime('%H:%M:%S')}找到图片 {image_name} 位置：{location}")
                                
                                # 执行原有的图片操作逻辑
                                current_x, current_y = location.x, location.y
                                x1, y1 = image_location.left, image_location.top
                                x2, y2 = x1 + image_location.width, y1 + image_location.height
                                
                                # 处理操作指令
                                ops = operations.split('，')
                                for op in ops:
                                    if self.stop_signal:
                                        break
                                    if op.startswith('偏移='):
                                        offset = op.split('=')[1].split('/')
                                        x_offset, y_offset = int(offset[0]), int(offset[1])
                                        pyautogui.move(x_offset, y_offset)
                                        current_x += x_offset
                                        current_y += y_offset
                                        self.log_message(f"{datetime.datetime.now().strftime('%H:%M:%S')}鼠标相对移动到偏移位置：({x_offset}, {y_offset})，当前坐标：({current_x}, {current_y})")
                                    elif op.startswith('等待='):
                                        wait_time = float(op.split('=')[1])
                                        time.sleep(wait_time)
                                        self.log_message(f"{datetime.datetime.now().strftime('%H:%M:%S')}等待 {wait_time} 秒")
                                    elif op == '左键按下':
                                        pyautogui.mouseDown(button='left')
                                        self.log_message(f"{datetime.datetime.now().strftime('%H:%M:%S')}左键按下")
                                    elif op == '左键释放':
                                        pyautogui.mouseUp(button='left')
                                        self.log_message(f"{datetime.datetime.now().strftime('%H:%M:%S')}左键释放")
                                    elif op == '右键按下':
                                        pyautogui.mouseDown(button='right')
                                        self.log_message(f"{datetime.datetime.now().strftime('%H:%M:%S')}右键按下")
                                    elif op == '右键释放':
                                        pyautogui.mouseUp(button='right')
                                        self.log_message(f"{datetime.datetime.now().strftime('%H:%M:%S')}右键释放")
                                    elif op.startswith('左键='):
                                        num_clicks = int(op.split('=')[1])
                                        for _ in range(num_clicks):
                                            click(current_x, current_y)
                                            self.log_message(f"{datetime.datetime.now().strftime('%H:%M:%S')}左键单击 {num_clicks} 次")
                                            if self.stop_signal:
                                                break
                                    elif op.startswith('二级左键='):
                                        num_clicks = int(op.split('=')[1])
                                        for _ in range(num_clicks):
                                            Shiji_x = generate_random_position(current_x, x1, x2, 1)
                                            Shiji_y = generate_random_position(current_y, y1, y2, 1)
                                            click(Shiji_x, Shiji_y)
                                            self.log_message(f"{datetime.datetime.now().strftime('%H:%M:%S')}随机偏移后单击 {num_clicks} 次")
                                            if self.stop_signal:
                                                break
                                    elif op.startswith('三级左键='):
                                        coords = op.split('=')[1].split('/')
                                        x_offset = int(coords[0])
                                        y_offset = int(coords[1])
                                        z_clicks = int(coords[2])
                                        for _ in range(z_clicks):
                                            Shiji_x = current_x + x_offset
                                            Shiji_y = current_y + y_offset
                                            click(Shiji_x, Shiji_y)
                                            self.log_message(
                                                f"{datetime.datetime.now().strftime('%H:%M:%S')}移动到 ({Shiji_x}, {Shiji_y}) 后左键点击 {z_clicks} 次")
                                            if self.stop_signal:
                                                break
                                    elif op.startswith('四级左键='):
                                        coords = op.split('=')[1].split('/')
                                        x_offset = int(coords[0])
                                        y_offset = int(coords[1])
                                        z_clicks = int(coords[2])
                                        for _ in range(z_clicks):
                                            Shiji_x = generate_random_position(current_x + x_offset, x1+ x_offset, x2+ x_offset, 1)
                                            Shiji_y = generate_random_position(current_y + y_offset, y1+ y_offset, y2+ y_offset, 1)
                                            click(Shiji_x, Shiji_y)
                                            self.log_message(
                                                f"{datetime.datetime.now().strftime('%H:%M:%S')}移动到 ({Shiji_x}, {Shiji_y}) 后左键点击 {z_clicks} 次")
                                            if self.stop_signal:
                                                break
                                    elif op == '屏幕截图':
                                        s_h = pyautogui.screenshot()
                                        screenshot_name = f"{datetime.datetime.now().strftime('%Y_%m_%d_%H_%M_%S.%f')}screenshot.png"
                                        # 使用选择的保存路径
                                        file_path = os.path.join(self.screenshot_path if self.screenshot_path else "lin/", screenshot_name)
                                        # 确保目录存在
                                        os.makedirs(os.path.dirname(file_path), exist_ok=True)
                                        s_h.save(file_path)
                                        self.log_message(f"{datetime.datetime.now().strftime('%H:%M:%S')}屏幕截图保存为 {screenshot_name}")
                                    elif op.startswith('区域屏幕截图='):
                                        region = tuple(map(int, op.split('=')[1].split('/')))
                                        region_screenshot_path = os.path.join(
                                            self.screenshot_path if self.screenshot_path else "lin/",
                                            f'region_screenshot_{datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S.%f")}.png'
                                        )
                                        # 确保目录存在
                                        os.makedirs(os.path.dirname(region_screenshot_path), exist_ok=True)
                                        pyautogui.screenshot(region_screenshot_path, region=region)
                                        self.log_message(f"区域屏幕截图保存为 {os.path.basename(region_screenshot_path)}, 区域：{region}")
                                    else:
                                        self.log_message(f"{datetime.datetime.now().strftime('%H:%M:%S')}未知操作：{op}")
                                    pass
                                
                                break
                        except Exception as e:
                            self.log_message(f"{datetime.datetime.now().strftime('%H:%M:%S')}查找图片时出错：{str(e)}")
                            time.sleep(0.1)  # 短暂延迟后重试

                    # 处理超时情况
                    if not image_found:
                        self.log_message(
                            f"{datetime.datetime.now().strftime('%H:%M:%S')}图片 {image_name} 查找超时")
                        
                        if timeout_action:
                            if timeout_action.lower() == 'skip':
                                self.log_message("跳过当前行")
                                current_row += 1
                            elif str(timeout_action).isdigit():
                                target_row = int(timeout_action)
                                self.log_message(f"跳转到第 {target_row} 行")
                                current_row = target_row
                            else:
                                current_row += 1
                        else:
                            current_row += 1
                    else:
                        current_row += 1

                    # 检查是否到达表格末尾
                    if current_row > max_row:
                        current_row = 2  # 回到开始行

                except Exception as e:
                    error_msg = f"主人，出错了喵~~~\n"
                    error_msg += f"在第 {current_row} 行发生了问题喵:\n"
                    error_msg += f"错误类型: {type(e).__name__}\n"
                    error_msg += f"错误描述: {str(e)}\n"
                    error_msg += f"当前处理的图片: {image_name if 'image_name' in locals() else '未知'}\n"
                    error_msg += f"当前的操作: {operations if 'operations' in locals() else '未知'}\n"
                    
                    self.log_error(error_msg)
                    self.log_message(f"主人，出错了喵~ 已经记录到日志文件了，错误信息: {str(e)}")
                    
                    # 继续处理下一行
                    current_row += 1
                    if current_row > max_row:
                        current_row = 2
                    
                    time.sleep(1)  # 出错后稍微暂停一下
                    continue
                
        except Exception as e:
            # 处理主循环的错误
            error_msg = "主人，程序出现大问题了喵~~~\n"
            error_msg += f"错误类型: {type(e).__name__}\n"
            error_msg += f"错误描述: {str(e)}\n"
            
            self.log_error(error_msg)
            self.log_message("主人，程序遇到严重错误了喵，请查看错误日志文件~")
        
        finally:
            self.log_message(f"{datetime.datetime.now().strftime('%H:%M:%S')} 程序运行结束了喵~")
            self.is_running = False

    def key_listener(self, event):
        if event.state == 4 and event.keysym == 's':
            self.stop()

    def load_settings(self):
        """从JSON文件加载所有设置"""
        try:
            print("开始加载设置...")  # 使用print替代log_error
            
            # 检查配置文件是否存在
            if not os.path.exists(self.CONFIG_FILE):
                print(f"配置文件不存在: {self.CONFIG_FILE}")  # 使用print
                return
            
            # 读取配置文件
            try:
                with open(self.CONFIG_FILE, 'r', encoding='utf-8') as f:
                    file_content = f.read()
                    if not file_content.strip():  # 检查文件是否为空
                        print("配置文件为空")  # 使用print
                        return
                    
                    loaded_settings = json.loads(file_content)
                    
                    # 应用设置
                    if 'confidence' in loaded_settings:
                        self.DEFAULT_CONFIDENCE = float(loaded_settings['confidence'])
                    if 'run_time' in loaded_settings:
                        self.DEFAULT_RUN_TIME = int(loaded_settings['run_time'])
                        if hasattr(self, 'custom_time_entry'):
                            self.custom_time_entry.delete(0, tk.END)
                            self.custom_time_entry.insert(0, str(self.DEFAULT_RUN_TIME))
                    if 'custom_time' in loaded_settings:
                        self.DEFAULT_CUSTOM_TIME = loaded_settings['custom_time']
                    if 'hotkey' in loaded_settings:
                        self.hotkey = set(loaded_settings['hotkey'])
                    if 'excel_path' in loaded_settings:
                        self.excel_path = str(loaded_settings['excel_path'])
                    if 'screenshot_path' in loaded_settings:
                        self.screenshot_path = str(loaded_settings['screenshot_path'])
                    
                    print("设置已成功加载")  # 使用print
                    
            except json.JSONDecodeError as e:
                self.log_error(f"JSON解析错误: {str(e)}\n文件内容: {file_content}")
            except Exception as e:
                self.log_error(f"读取配置文件时出错: {str(e)}")
            
        except Exception as e:
            self.log_error(f"加载设置过程中出现未知错误: {str(e)}\n{self.get_traceback()}")

    def save_settings(self):
        """保存所有设置到JSON文件"""
        try:
            # 获取当前UI中的值
            current_settings = {
                'confidence': float(self.confidence_entry.get()) if self.confidence_entry.get() else self.DEFAULT_CONFIDENCE,
                'run_time': int(self.custom_time_entry.get()) if self.custom_time_entry.get() else self.DEFAULT_RUN_TIME,
                'custom_time': self.DEFAULT_CUSTOM_TIME,
                'hotkey': list(self.hotkey) if hasattr(self, 'hotkey') else [],
                'excel_path': self.excel_entry.get() if hasattr(self, 'excel_entry') else '',
                'screenshot_path': self.screenshot_entry.get() if hasattr(self, 'screenshot_entry') else ''
            }
            
            # 清空并重写配置文件
            with open(self.CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(current_settings, f, indent=4, ensure_ascii=False)
            
            print(f"设置已保存到: {self.CONFIG_FILE}")  # 使用print
            
        except Exception as e:
            self.log_error(f"保存设置时出错: {str(e)}\n{self.get_traceback()}")

    def on_confidence_change(self, event=None):
        """当精度值改变时保存配置"""
        try:
            confidence = float(self.confidence_entry.get())
            if 0.1 <= confidence <= 1.0:
                self.DEFAULT_CONFIDENCE = confidence
                self.save_settings()
                self.log_message(f"已保存新的精度值: {confidence}")
        except ValueError:
            self.log_message("无效的精度值")
            # 恢复为上一次的有效值
            self.confidence_entry.delete(0, tk.END)
            self.confidence_entry.insert(0, str(self.DEFAULT_CONFIDENCE))

    def log_error(self, error_message, is_error=True):
        """记录错误到文件
        Args:
            error_message: 错误信息
            is_error: 是否为错误信息，默认为True。只有错误信息才会写入日志文件。
        """
        try:
            # 如果不是错误信息，只打印到控制台
            if not is_error:
                print(error_message)
                return
            
            # 只有实际的错误才写入日志文件
            current_time = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            error_file = os.path.join(self.log_folder, f'error_log_{current_time}.txt')
            
            with open(error_file, 'a', encoding='utf-8') as f:
                f.write(f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]\n")
                f.write(f"{error_message}\n")
                f.write("-" * 50 + "\n")
            
            print(f"错误已记录到: {error_file}")
        except Exception as e:
            print(f"记录错误日志时出现问题: {str(e)}")

    def get_traceback(self):
        """获取完整的错误追踪信息"""
        import traceback
        return traceback.format_exc()

    def check_program_validity(self):
        """检查程序是否在有效期内"""
        expiry_date = datetime.datetime(2128, 1, 17)  # 设置过期时间
        current_date = datetime.datetime.now()
        
        if current_date.date() >= expiry_date.date():
            import tkinter.messagebox as messagebox
            messagebox.showerror(
                "程序已过期", 
                "主人喵~ 这个版本已经过期了呢！\n"
                "请使用新版本继续使用喵~\n"
                "如需帮助请联系作者水年喵~"
            )
            return False
        return True

    def load_hotkey(self):
        try:
            with open(self.CONFIG_FILE, 'r') as f:
                config = json.load(f)
                self.hotkey = set(config.get('hotkey', []))
                self.update_hotkey_display()
        except FileNotFoundError:
            self.hotkey = set()

    def save_hotkey(self):
        with open(self.CONFIG_FILE, 'w') as f:
            json.dump({'hotkey': list(self.hotkey)}, f)

    def update_hotkey_display(self):
        """更新快捷键显示"""
        self.hotkey_display.configure(state='normal')
        self.hotkey_display.delete(0, tk.END)
        # 按照特定顺序显示组合键（修饰键在前）
        sorted_keys = sorted(self.hotkey, key=lambda x: (x not in {'CTRL', 'ALT', 'SHIFT'}, x))
        self.hotkey_display.insert(0, ' + '.join(sorted_keys))
        self.hotkey_display.configure(state='readonly')

    def on_press(self, key):
        """处理按键按下事件"""
        try:
            key_str = self.get_key_string(key)
            
            if self.is_setting_hotkey:
                self.temp_hotkey.add(key_str)
                if len(self.temp_hotkey) > 2:
                    self.temp_hotkey.clear()
                self.update_hotkey_display_temp()
            else:
                self.current_keys.add(key_str)
                # 检查当前按下的键是否匹配保存的快捷键
                if self.hotkey and self.current_keys == self.hotkey:
                    self.stop_program()
        except AttributeError:
            pass

    def get_key_string(self, key):
        """获取按键的标准显示字符串"""
        try:
            if hasattr(key, 'char'):
                if key.char:  # 处理普通字符键
                    return key.char.upper()  # 转换为大写以统一显示
                else:  # 处理控制键组合
                    return str(key).replace('Key.', '').upper()
            else:  # 处理特殊键
                special_keys = {
                    'Key.ctrl_l': 'CTRL',
                    'Key.ctrl_r': 'CTRL',
                    'Key.alt_l': 'ALT',
                    'Key.alt_r': 'ALT',
                    'Key.shift_l': 'SHIFT',
                    'Key.shift_r': 'SHIFT',
                    'Key.space': 'SPACE',
                    'Key.enter': 'ENTER',
                    'Key.esc': 'ESC'
                }
                key_str = str(key)
                return special_keys.get(key_str, key_str.replace('Key.', '').upper())
        except:
            return str(key).replace('Key.', '').upper()

    def on_release(self, key):
        """处理按键释放事件"""
        try:
            if hasattr(key, 'char'):
                key_str = key.char
            else:
                key_str = str(key).replace('Key.', '')
            
            if key_str in self.current_keys:
                self.current_keys.remove(key_str)
        except AttributeError:
            pass

        if self.is_setting_hotkey and not self.current_keys:
            self.finish_hotkey_setting()

    def start_hotkey_setting(self):
        self.is_setting_hotkey = True
        self.temp_hotkey.clear()
        self.current_keys.clear()
        
        # 停止现有的监听器
        if self.listener:
            self.listener.stop()
        
        # 创建新的监听器
        self.listener = keyboard.Listener(on_press=self.on_press, on_release=self.on_release)
        self.listener.start()
        
        # 创建确认按钮的弹窗
        self.confirm_window = tk.Toplevel(self.root)
        self.confirm_window.title("设置快捷键")
        self.confirm_window.geometry("300x150")
        
        label = ttk.Label(self.confirm_window, text="请按下键盘按键\n(最多支持两个组合键)")
        label.pack(pady=20)
        
        confirm_button = ttk.Button(self.confirm_window, text="确认", command=self.finish_hotkey_setting)
        confirm_button.pack(pady=10)
        
        self.confirm_window.transient(self.root)
        self.confirm_window.grab_set()

    def finish_hotkey_setting(self):
        # 停止当前的监听器
        if self.listener:
            self.listener.stop()
        
        self.is_setting_hotkey = False
        if self.temp_hotkey:
            self.hotkey = self.temp_hotkey.copy()
            self.save_hotkey()
        
        self.temp_hotkey.clear()
        self.current_keys.clear()
        self.update_hotkey_display()
        
        if hasattr(self, 'confirm_window'):
            self.confirm_window.destroy()
        
        # 重新启动新的监听器用于常规监听
        self.listener = keyboard.Listener(on_press=self.on_press, on_release=self.on_release)
        self.listener.start()

    def update_hotkey_display_temp(self):
        """更新临时快捷键显示"""
        self.hotkey_display.configure(state='normal')
        self.hotkey_display.delete(0, tk.END)
        # 按照特定顺序显示组合键（修饰键在前）
        sorted_keys = sorted(self.temp_hotkey, key=lambda x: (x not in {'CTRL', 'ALT', 'SHIFT'}, x))
        self.hotkey_display.insert(0, ' + '.join(sorted_keys))
        self.hotkey_display.configure(state='readonly')

    def stop_program(self):
        """通过快捷键停止程序"""
        if self.is_running:
            self.stop()  # 调用现有的stop方法
            self.log_message("通过快捷键停止程序")
            messagebox.showinfo("程序停止", "程序已通过快捷键停止")

    def start_hotkey_listener(self):
        """启动快捷键监听器"""
        # 确保之前的监听器已停止
        if hasattr(self, 'listener') and self.listener:
            self.listener.stop()
        
        # 创建并启动新的监听器
        self.listener = keyboard.Listener(on_press=self.on_press, on_release=self.on_release)
        self.listener.daemon = True  # 设置为守护线程
        self.listener.start()
        
        # 加载保存的快捷键
        self.load_hotkey()

if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = AutoScriptApp(root)
        # only程序有效时运行主循环
        if hasattr(app, 'root'):  # 检查初始化
            root.bind('<KeyPress>', app.key_listener)
            root.mainloop()
    except Exception as e:
        import tkinter.messagebox as messagebox
        messagebox.showerror("错误", f"主人喵~ 程序出现问题了：\n{str(e)}")