"""读取 RPA 配置、Excel 任务和模型类别。"""
from dataclasses import dataclass
import json
import os
from pathlib import Path
import threading
from typing import Any

import openpyxl
import yaml


_MODEL_CACHE: dict[Path, Any] = {}
_MODEL_CACHE_LOCK = threading.Lock()


@dataclass(frozen=True)
class AppConfig:
    root: Path
    confidence: float = 0.8
    excel_path: str = ""
    sheet_name: str = ""
    screenshot_path: str = ""
    hotkey: tuple[str, ...] = ()
    model_path: str = "best.pt"
    class_names: tuple[str, ...] = ()
    class_name_to_chinese: dict[str, str] | None = None
    actions: dict[str, dict[str, Any]] | None = None


def project_root() -> Path:
    if getattr(__import__("sys"), "frozen", False):
        return Path(getattr(__import__("sys"), "_MEIPASS"))
    return Path(__file__).resolve().parent


def _load_yaml(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as stream:
        return yaml.safe_load(stream) or {}


def _resolve_model_path(root: Path, model_path: str | Path) -> Path:
    candidate = Path(os.path.expandvars(os.path.expanduser(str(model_path))))
    if candidate.is_absolute():
        return candidate
    return (root / candidate).resolve()


def _normalize_class_names(raw_names: Any) -> tuple[str, ...]:
    if isinstance(raw_names, dict):
        ordered_names = [raw_names[index] for index in sorted(raw_names, key=lambda value: int(str(value)))]
        return tuple(str(name) for name in ordered_names)
    if isinstance(raw_names, list):
        return tuple(str(name) for name in raw_names)
    return ()


def load_model(model_path: str | Path, root: Path | None = None) -> Any:
    root = root or project_root()
    model_file = _resolve_model_path(root, model_path)
    with _MODEL_CACHE_LOCK:
        cached_model = _MODEL_CACHE.get(model_file)
        if cached_model is not None:
            return cached_model
        import torch
        loaded_model = torch.hub.load(str(root), "custom", path=str(model_file), source="local")
        _MODEL_CACHE[model_file] = loaded_model
        return loaded_model


def read_model_metadata(model_path: str | Path, root: Path | None = None) -> tuple[tuple[str, ...], dict[str, str]]:
    root = root or project_root()
    model_file = _resolve_model_path(root, model_path)
    candidates: list[Path] = []
    if model_file.exists():
        candidates.extend([
            model_file.with_suffix(".yaml"),
            model_file.with_suffix(".yml"),
        ])
    if model_file.name.lower().endswith((".pt", ".pth")):
        base_name = model_file.name.rsplit(".", 1)[0]
        candidates.extend([
            model_file.with_name(f"{base_name}.yaml"),
            model_file.with_name(f"{base_name}.yml"),
        ])
    for candidate in candidates:
        if not candidate.exists():
            continue
        metadata = _load_yaml(candidate)
        if not isinstance(metadata, dict):
            continue
        names = metadata.get("names", metadata.get("class_names", []))
        class_names = _normalize_class_names(names)
        if class_names:
            translations = metadata.get("Class_Name_To_Chinese", {})
            if not isinstance(translations, dict):
                translations = {}
            return class_names, {str(key): str(value) for key, value in translations.items()}

    try:
        loaded_model = load_model(model_file, root)
        names = getattr(loaded_model, "names", None)
        if isinstance(names, dict):
            ordered_names = [names[index] for index in sorted(names, key=lambda value: int(str(value)))]
            return tuple(str(name) for name in ordered_names), {}
        if isinstance(names, (list, tuple)):
            return tuple(str(name) for name in names), {}
    except Exception:
        pass

    default_names = _load_yaml(root / "data" / "onmyoji.yaml").get("names", [])
    default_translations = _load_yaml(root / "data" / "onmyoji_name.yaml").get("Class_Name_To_Chinese", {})
    return _normalize_class_names(default_names), {str(key): str(value) for key, value in default_translations.items()}


def read_app_config(root: Path | None = None, settings_path: Path | None = None) -> AppConfig:
    root = root or project_root()
    settings_path = settings_path or root / "settings.json"
    settings: dict[str, Any] = {}
    if settings_path.exists():
        try:
            settings = json.loads(settings_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            settings = {}

    try:
        confidence = float(settings.get("confidence", 0.8))
    except (TypeError, ValueError):
        confidence = 0.8
    confidence = min(1.0, max(0.1, confidence))

    model_path = str(settings.get("model_path", "best.pt"))
    class_names, translations = read_model_metadata(model_path, root)
    actions = _load_yaml(root / "data" / "3leader.yaml").get("actions", {})
    hotkey = tuple(str(value) for value in settings.get("hotkey", []))
    return AppConfig(
        root=root,
        confidence=confidence,
        excel_path=str(settings.get("excel_path", "")),
        sheet_name=str(settings.get("sheet_name", "")),
        screenshot_path=str(settings.get("screenshot_path", "")),
        hotkey=hotkey,
        model_path=model_path,
        class_names=class_names,
        class_name_to_chinese=translations,
        actions=actions,
    )


def save_settings(config: AppConfig, path: Path | None = None) -> None:
    path = path or config.root / "settings.json"
    data = {
        "confidence": config.confidence,
        "excel_path": config.excel_path,
        "sheet_name": config.sheet_name,
        "screenshot_path": config.screenshot_path,
        "hotkey": list(config.hotkey),
        "model_path": config.model_path,
    }
    path.write_text(json.dumps(data, ensure_ascii=False, indent=4), encoding="utf-8")


def read_excel_tasks(path: str | Path, sheet_name: str) -> list[tuple[Any, Any, Any, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表不存在: {sheet_name}")
        sheet = workbook[sheet_name]
        return [tuple(row) for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True)]
    finally:
        workbook.close()


def resolve_image_path(image_name: str, root: Path) -> Path:
    candidate = Path(os.path.expandvars(os.path.expanduser(image_name)))
    if candidate.is_absolute():
        return candidate
    return root / candidate


def validate_startup_inputs(config: AppConfig) -> None:
    if not config.model_path:
        raise ValueError("请选择模型文件")

    model_path = _resolve_model_path(config.root, config.model_path)
    if not model_path.exists():
        raise FileNotFoundError(f"模型文件不存在: {model_path}")

    if not config.class_names:
        raise ValueError("模型类别为空，请确认模型文件和同名 YAML 是否正确")

    action_names = set((config.actions or {}).keys())
    missing_action_names = sorted(name for name in action_names if name not in config.class_names)
    if missing_action_names:
        joined_names = ", ".join(missing_action_names[:10])
        suffix = "..." if len(missing_action_names) > 10 else ""
        raise ValueError(
            "模型类别不匹配：动作配置里存在未定义类别 "
            f"{joined_names}{suffix}。请确认模型是否和目标类一致。"
        )

    if not config.excel_path:
        raise ValueError("请选择 Excel 文件")
    excel_path = Path(os.path.expandvars(os.path.expanduser(config.excel_path)))
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

    if not config.sheet_name:
        raise ValueError("请选择工作表")

    tasks = read_excel_tasks(excel_path, config.sheet_name)
    bad_images: list[str] = []
    valid_class_names = {str(name).strip().casefold() for name in config.class_names}
    for row_index, (image_name, _, _, _) in enumerate(tasks, start=2):
        if not image_name:
            continue
        image_name_text = str(image_name).strip()
        lowered = image_name_text.casefold()
        if lowered.endswith((".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp", ".tif", ".tiff")) or "/" in image_name_text or "\\" in image_name_text:
            image_path = resolve_image_path(image_name_text, config.root)
            if not image_path.exists():
                bad_images.append(f"第 {row_index} 行: {image_name}")
            continue
        if lowered in valid_class_names:
            continue
        image_path = resolve_image_path(image_name_text, config.root)
        if not image_path.exists():
            bad_images.append(f"第 {row_index} 行: {image_name}")
    if bad_images:
        raise ValueError(
            "Excel 中以下图片路径或目标类不存在，请检查文件名或模型类别：\n"
            + "\n".join(bad_images[:10])
            + ("\n..." if len(bad_images) > 10 else "")
        )
